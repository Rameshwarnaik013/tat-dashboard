import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
from datetime import timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="TAT Dashboard", layout="wide")

st.markdown("""
<style>
    .block-container { padding-top: 1rem; }
    [data-testid="stMetricValue"] { font-size: 1.2rem; }
    .stTabs [data-baseweb="tab"] { font-size: 1rem; font-weight: 600; }
    section[data-testid="stSidebar"] .stExpander {
        border: 1px solid #ddd; border-radius: 8px; margin-bottom: 0.35rem;
    }
</style>
""", unsafe_allow_html=True)

DISPATCH_CATS = ["0-2 Days", "3-5 Days", ">5 Days"]
DELIVERY_CATS = ["0-9 Days", "10-12 Days", ">12 Days", "Delivery Pending"]

DISPATCH_COLORS = {
    "0-2 Days": "#27ae60", "3-5 Days": "#f39c12", ">5 Days": "#e74c3c",
}
DELIVERY_COLORS = {
    "0-9 Days": "#27ae60", "10-12 Days": "#f39c12",
    ">12 Days": "#e74c3c", "Delivery Pending": "#7f8c8d",
}

_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_HDR_FILL = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center")
_GRAND_FONT = Font(bold=True, size=11)
_GRAND_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center")


def load_file(file):
    if file.name.endswith(".csv"):
        return pd.read_csv(file)
    return pd.read_excel(file)


def derive_dispatch_tat(df):
    if "Dis Days" not in df.columns:
        return
    if "Dis TAT" not in df.columns:
        df["Dis TAT"] = np.nan
    empty = df["Dis TAT"].isna() | (df["Dis TAT"].astype(str).str.strip() == "")
    if not empty.any():
        return
    days = pd.to_numeric(df.loc[empty, "Dis Days"], errors="coerce")
    df.loc[empty, "Dis TAT"] = np.where(
        days <= 2, "0-2 Days", np.where(days <= 5, "3-5 Days", ">5 Days")
    )


def derive_delivery_tat(df):
    if "Delivery Days" not in df.columns:
        return
    if "Delivery TAT" not in df.columns:
        df["Delivery TAT"] = np.nan
    empty = df["Delivery TAT"].isna() | (df["Delivery TAT"].astype(str).str.strip() == "")
    if not empty.any():
        return
    days = pd.to_numeric(df.loc[empty, "Delivery Days"], errors="coerce")
    conditions = [days.isna(), days <= 9, days <= 12, days > 12]
    choices = ["Delivery Pending", "0-9 Days", "10-12 Days", ">12 Days"]
    df.loc[empty, "Delivery TAT"] = np.select(conditions, choices, default="Delivery Pending")


def build_pivot(data, group_col, tat_col, categories, date_sort=False):
    valid = data[data[tat_col].notna() & (data[tat_col].astype(str).str.strip() != "")]
    if valid.empty:
        empty = pd.DataFrame(columns=categories, index=["Grand Total"]).fillna("0%")
        return empty, pd.DataFrame(columns=categories, index=["Grand Total"]).fillna(0)

    ct = pd.crosstab(valid[group_col], valid[tat_col])
    for cat in categories:
        if cat not in ct.columns:
            ct[cat] = 0
    ct = ct[categories]
    if date_sort:
        ct = ct.sort_index()

    row_sums = ct.sum(axis=1).replace(0, np.nan)
    pct = ct.div(row_sums, axis=0).mul(100).fillna(0).round(0).astype(int)

    grand = ct.sum(axis=0)
    gt = grand.sum()
    if gt > 0:
        grand_pct = (grand / gt * 100).round(0).astype(int)
    else:
        grand_pct = (grand * 0).astype(int)
    grand_pct.name = "Grand Total"

    pct = pd.concat([pct, grand_pct.to_frame().T])
    formatted = pct.astype(str) + "%"
    return formatted, pct


def style_pivot(fmt_df):
    def _bg(val):
        try:
            v = int(str(val).replace("%", ""))
        except (ValueError, TypeError):
            return ""
        if v >= 50:
            return "background-color: #f5b7b1"
        if v >= 30:
            return "background-color: #fdebd0"
        return "background-color: #d5f5e3"

    styler = fmt_df.style.map(_bg)

    styler = styler.set_table_styles([
        {"selector": "th",
         "props": [("background-color", "#1a5276"), ("color", "white"),
                   ("font-weight", "bold"), ("text-align", "center"),
                   ("padding", "6px 12px")]},
        {"selector": "th.row_heading",
         "props": [("background-color", "#1a5276"), ("color", "white"),
                   ("font-weight", "bold"), ("text-align", "left")]},
    ])

    if "Grand Total" in fmt_df.index:
        styler = styler.set_properties(
            subset=pd.IndexSlice["Grand Total", :],
            **{"font-weight": "bold", "background-color": "#d6eaf8"},
        )
    return styler


def render_chart(pct_df, title, categories, colors):
    chart = pct_df.drop("Grand Total", errors="ignore")
    if chart.empty:
        return go.Figure().update_layout(title=title, height=300)
    fig = go.Figure()
    for cat in categories:
        if cat in chart.columns:
            fig.add_trace(
                go.Bar(
                    name=cat,
                    x=chart.index.astype(str),
                    y=chart[cat],
                    marker_color=colors.get(cat, "#95a5a6"),
                    text=chart[cat].astype(str) + "%",
                    textposition="inside",
                )
            )
    fig.update_layout(
        barmode="stack", title=title, height=420,
        yaxis_title="Percentage", xaxis_title="",
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
        margin=dict(t=60, b=40),
    )
    return fig


def show_kpis(fdf, tat_col, categories):
    n_total = len(fdf)
    valid = fdf[tat_col].dropna()
    valid = valid[valid.astype(str).str.strip() != ""]
    n_valid = len(valid)
    tc = valid.value_counts()

    n_cols = 2 + len(categories)
    cols = st.columns(n_cols)
    cols[0].metric("Total Line Items", f"{n_total:,}")
    for i, cat in enumerate(categories, 1):
        cnt = tc.get(cat, 0)
        pct = cnt / n_valid * 100 if n_valid > 0 else 0
        cols[i].metric(cat, f"{cnt:,}  ({pct:.0f}%)")
    if "Sale Invoice" in fdf.columns:
        cols[-1].metric("Invoices", f"{fdf['Sale Invoice'].nunique():,}")


def render_tat_section(fdf, tat_col, tat_label, categories, colors, key_pfx):
    show_kpis(fdf, tat_col, categories)
    st.divider()

    sub_tabs = st.tabs([
        "Invoice Date", "MIS Item Group", "Sales Channel", "From Warehouse",
    ])
    views = [
        ("Invoice Date", "_date_str", True),
        ("MIS Item Group", "New MIS Item Group", False),
        ("Sales Channel", "Sales_Channel", False),
        ("From Warehouse", "From Warehouse", False),
    ]

    pivots = {}

    for sub_tab, (label, col, dsort) in zip(sub_tabs, views):
        with sub_tab:
            st.subheader(f"{tat_label} by {label}")
            fmt, pct = build_pivot(fdf, col, tat_col, categories, date_sort=dsort)
            pivots[label] = fmt

            st.dataframe(
                style_pivot(fmt), use_container_width=True,
                height=min(len(fmt) * 38 + 60, 700),
            )
            st.plotly_chart(
                render_chart(pct, f"{tat_label} - {label}", categories, colors),
                use_container_width=True,
            )

            if label == "Sales Channel":
                st.divider()
                st.subheader("Customer Breakdown within Sales Channel")

                channels = sorted(fdf["Sales_Channel"].dropna().unique().tolist())
                if not channels:
                    st.info("No channel data.")
                    continue

                chosen_ch = st.selectbox(
                    "Select Sales Channel",
                    channels,
                    key=f"{key_pfx}_ch_drill",
                )
                ch_data = fdf[fdf["Sales_Channel"] == chosen_ch]
                valid_ch = ch_data[tat_col].dropna()
                valid_ch = valid_ch[valid_ch.astype(str).str.strip() != ""]

                if valid_ch.empty:
                    st.info("No data for this channel.")
                    continue

                c_fmt, c_pct = build_pivot(ch_data, "Customer", tat_col, categories)
                pivots[f"Cust-{chosen_ch}"] = c_fmt

                st.dataframe(
                    style_pivot(c_fmt), use_container_width=True,
                    height=min(len(c_fmt) * 38 + 60, 500),
                )
                st.plotly_chart(
                    render_chart(c_pct, f"Customers in {chosen_ch}", categories, colors),
                    use_container_width=True,
                )

    return pivots


def smart_filter(label, options, key):
    sel_key = f"sel_{key}"

    if sel_key not in st.session_state:
        st.session_state[sel_key] = list(options)

    options_set = set(options)
    valid = [v for v in st.session_state[sel_key] if v in options_set]
    if len(valid) != len(st.session_state[sel_key]):
        st.session_state[sel_key] = valid

    n_sel = len(st.session_state.get(sel_key, []))
    n_total = len(options)
    count_txt = "All" if n_sel == n_total else f"{n_sel} of {n_total}"

    with st.sidebar.expander(f"{label} ({count_txt} selected)"):
        search = st.text_input(
            "search", key=f"srch_{key}",
            placeholder=f"Search {label.lower()}...",
            label_visibility="collapsed",
        )

        c1, c2 = st.columns(2)
        if c1.button("All", key=f"all_{key}", use_container_width=True):
            st.session_state[sel_key] = list(options)
            st.session_state[f"srch_{key}"] = ""
            st.rerun()
        if c2.button("None", key=f"clr_{key}", use_container_width=True):
            st.session_state[sel_key] = []
            st.session_state[f"srch_{key}"] = ""
            st.rerun()

        if search.strip():
            q = search.strip().lower()
            sel_set = set(st.session_state.get(sel_key, []))
            visible = [o for o in options if q in str(o).lower() or o in sel_set]
        else:
            visible = options

        selected = st.multiselect(
            label, visible, key=sel_key,
            placeholder="Select...",
            label_visibility="collapsed",
        )
        return selected


def _format_sheet(ws):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        cell.border = _THIN_BORDER

    for row in range(2, ws.max_row + 1):
        is_grand = str(ws.cell(row=row, column=1).value).strip() == "Grand Total"
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER if col > 1 else Alignment(horizontal="left")
            if is_grand:
                cell.font = _GRAND_FONT
                cell.fill = _GRAND_FILL

    for col in range(1, ws.max_column + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = max_len + 4


def build_excel(dispatch_pivots, delivery_pivots):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, p in dispatch_pivots.items():
            sheet = f"Dis-{name}"[:31]
            p.to_excel(writer, sheet_name=sheet)
            _format_sheet(writer.sheets[sheet])

        for name, p in delivery_pivots.items():
            sheet = f"Del-{name}"[:31]
            p.to_excel(writer, sheet_name=sheet)
            _format_sheet(writer.sheets[sheet])
    return buf.getvalue()


st.title("TAT Dashboard")
st.caption("Dispatch & Delivery Turnaround Time Analysis")

uploaded = st.file_uploader("Upload TAT Report", type=["csv", "xlsx", "xls"])
if not uploaded:
    st.info("Upload your sales order report (CSV or Excel) to begin.")
    st.stop()

df = load_file(uploaded)

REQUIRED = [
    "Invoice Date", "New MIS Item Group", "Sales_Channel",
    "From Warehouse", "Customer",
]
missing = [c for c in REQUIRED if c not in df.columns]
if missing:
    st.error(f"Missing columns: {', '.join(missing)}")
    st.stop()

df["Invoice Date"] = pd.to_datetime(df["Invoice Date"], errors="coerce")
df.dropna(subset=["Invoice Date"], inplace=True)
df["_date_str"] = df["Invoice Date"].dt.strftime("%Y-%m-%d")

if "Dis TAT" not in df.columns:
    df["Dis TAT"] = np.nan
if "Delivery TAT" not in df.columns:
    df["Delivery TAT"] = np.nan

derive_dispatch_tat(df)
derive_delivery_tat(df)


st.sidebar.header("Filters")

if st.sidebar.button("Reset All Filters", use_container_width=True, type="secondary"):
    for k in list(st.session_state.keys()):
        if k.startswith(("sel_", "all_", "clr_", "srch_", "date_")):
            del st.session_state[k]
    st.rerun()

st.sidebar.markdown("---")

min_d = df["Invoice Date"].min().date()
max_d = df["Invoice Date"].max().date()

with st.sidebar.expander("Invoice Date Range", expanded=True):
    preset = st.radio(
        "Quick select",
        ["All Time", "Last 7 Days", "Last 15 Days", "Last 30 Days", "Custom"],
        horizontal=True,
        key="date_preset",
        label_visibility="collapsed",
    )
    if preset == "Last 7 Days":
        start_d = max_d - timedelta(days=6)
        end_d = max_d
    elif preset == "Last 15 Days":
        start_d = max_d - timedelta(days=14)
        end_d = max_d
    elif preset == "Last 30 Days":
        start_d = max_d - timedelta(days=29)
        end_d = max_d
    elif preset == "All Time":
        start_d = min_d
        end_d = max_d
    else:
        d_range = st.date_input(
            "Pick range",
            value=(min_d, max_d),
            min_value=min_d,
            max_value=max_d,
            label_visibility="collapsed",
        )
        start_d = d_range[0] if isinstance(d_range, (list, tuple)) and len(d_range) >= 1 else min_d
        end_d = d_range[1] if isinstance(d_range, (list, tuple)) and len(d_range) >= 2 else max_d

    if start_d < min_d:
        start_d = min_d
    st.caption(f"{start_d.strftime('%d %b %Y')}  to  {end_d.strftime('%d %b %Y')}")

wh_opts = sorted(df["From Warehouse"].dropna().unique().tolist())
item_opts = sorted(df["New MIS Item Group"].dropna().unique().tolist())
ch_opts = sorted(df["Sales_Channel"].dropna().unique().tolist())

sel_wh = smart_filter("From Warehouse", wh_opts, "wh")
sel_item = smart_filter("MIS Item Group", item_opts, "item")
sel_ch = smart_filter("Sales Channel", ch_opts, "ch")

cust_pool = sorted(
    df.loc[df["Sales_Channel"].isin(sel_ch), "Customer"]
    .dropna().unique().tolist()
) if sel_ch else sorted(df["Customer"].dropna().unique().tolist())

sel_cust = smart_filter("Customer", cust_pool, "cust")

if not sel_wh or not sel_item or not sel_ch or not sel_cust:
    st.warning(
        "One or more filters have no selection. "
        "Use 'All' or pick at least one item in each filter."
    )
    st.stop()

mask = (
    (df["Invoice Date"].dt.date >= start_d)
    & (df["Invoice Date"].dt.date <= end_d)
    & (df["From Warehouse"].isin(sel_wh))
    & (df["New MIS Item Group"].isin(sel_item))
    & (df["Sales_Channel"].isin(sel_ch))
    & (df["Customer"].isin(sel_cust))
)
fdf = df.loc[mask].copy()

if fdf.empty:
    st.warning("No data matches the selected filters.")
    st.stop()

st.sidebar.markdown("---")
st.sidebar.metric("Filtered Rows", f"{len(fdf):,}")
st.sidebar.caption(
    f"WH {len(sel_wh)}/{len(wh_opts)}  |  "
    f"Item {len(sel_item)}/{len(item_opts)}  |  "
    f"Ch {len(sel_ch)}/{len(ch_opts)}  |  "
    f"Cust {len(sel_cust)}/{len(cust_pool)}"
)


dispatch_tab, delivery_tab = st.tabs(["DISPATCH TAT", "DELIVERY TAT"])

with dispatch_tab:
    dispatch_pivots = render_tat_section(
        fdf, "Dis TAT", "Dispatch TAT", DISPATCH_CATS, DISPATCH_COLORS, "dis",
    )

with delivery_tab:
    delivery_pivots = render_tat_section(
        fdf, "Delivery TAT", "Delivery TAT", DELIVERY_CATS, DELIVERY_COLORS, "del",
    )


st.divider()
st.subheader("Export Report")

st.download_button(
    "Download TAT Pivot Report (Excel)",
    build_excel(dispatch_pivots, delivery_pivots),
    "TAT_Pivot_Report.xlsx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
